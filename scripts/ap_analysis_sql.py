"""
AP Ad-Hoc Analysis Queries
Author: Augusto Sander

Purpose:
SQL-based analysis layer for answering business questions
outside the standard dashboard reporting cycle.

Business context:
The dashboard backend (ap_dashboard_backend.py) handles
recurring monthly KPI reporting. This script handles the
ad-hoc questions a controller, AP manager, or CFO would
ask when they need a specific answer fast — without
rebuilding the entire reporting pipeline.

How it works:
- Loads the dashboard-ready CSV into an in-memory SQLite database
- Validates for duplicate invoice numbers before any query runs
- Runs 8 named SQL queries against the data
- Exports all results to a formatted Excel file — one tab per question

How to run:
    python scripts/ap_analysis_sql.py

Requirement:
    Run ap_dashboard_backend.py first to generate:
    output/ap_invoices_dashboard_ready.csv

Output:
    output/ap_analysis_results.xlsx
    One tab per business question, ready to share with management.
"""

import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# File paths
INPUT_FILE = "output/ap_invoices_dashboard_ready.csv"
OUTPUT_FILE = "output/ap_analysis_results.xlsx"

# Color palette — matches the dashboard
NAVY = "1F3864"
STEEL_BLUE = "2E75B6"
GOLD = "C9A84C"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
RED = "C00000"
AMBER = "ED7D31"
GREEN = "375623"


# ─────────────────────────────────────────────
# DATA LOADING AND VALIDATION
# ─────────────────────────────────────────────

def load_data_to_sqlite(file_path):
    """
    Load the dashboard-ready CSV into an in-memory SQLite database.
    Validates for duplicate invoice numbers before loading.

    Business meaning:
    Duplicate invoice numbers in a real ERP export indicate either
    a data quality issue or a genuine duplicate payment risk.
    Catching this before analysis ensures query results are reliable.
    """
    df = pd.read_csv(file_path)

    # Validate: check for duplicate invoice numbers
    duplicates = df[df.duplicated(subset=["invoice_num"], keep=False)]
    if not duplicates.empty:
        dup_list = duplicates["invoice_num"].unique().tolist()
        print(f"  WARNING: Duplicate invoice numbers found: {dup_list}")
        print(f"  These rows may affect query accuracy.")
    else:
        print(f"  Validation passed: no duplicate invoice numbers found.")

    conn = sqlite3.connect(":memory:")
    df.to_sql("ap_data", conn, index=False, if_exists="replace")

    return conn, df


# ─────────────────────────────────────────────
# SQL QUERIES — 8 BUSINESS QUESTIONS
# ─────────────────────────────────────────────

def query_late_payment_frequency(conn):
    """
    Q1: Which vendors have the highest late payment frequency?

    Business meaning:
    Late payments damage vendor relationships and can trigger
    penalty clauses or supply disruptions. Identifying which
    vendors are consistently paid late helps AP management
    prioritize process improvements and relationship repair.
    """
    sql = """
        SELECT
            vendor_name                                                        AS "Vendor",
            COUNT(*)                                                           AS "Total Paid",
            SUM(CASE WHEN late_payment_flag = 1 THEN 1 ELSE 0 END)            AS "Paid Late",
            ROUND(
                100.0 * SUM(CASE WHEN late_payment_flag = 1 THEN 1 ELSE 0 END)
                / COUNT(*), 1
            )                                                                  AS "Late Rate %",
            ROUND(AVG(days_to_pay), 1)                                         AS "Avg Days to Pay"
        FROM ap_data
        WHERE status = 'Paid'
        GROUP BY vendor_name
        HAVING "Paid Late" > 0
        ORDER BY "Late Rate %" DESC
    """
    return pd.read_sql_query(sql, conn)


def query_avg_days_to_pay_by_expense(conn):
    """
    Q2: What is the average days to pay by expense type?

    Business meaning:
    Different expense categories carry different payment urgency.
    Freight and inventory vendors often have tighter windows than
    facilities or office supply vendors. This query shows whether
    AP is treating high-priority categories with appropriate speed.
    """
    sql = """
        SELECT
            expense_type                        AS "Expense Type",
            COUNT(*)                            AS "Paid Invoices",
            ROUND(AVG(days_to_pay), 1)          AS "Avg Days to Pay",
            MIN(days_to_pay)                    AS "Fastest",
            MAX(days_to_pay)                    AS "Slowest"
        FROM ap_data
        WHERE status = 'Paid'
            AND days_to_pay IS NOT NULL
        GROUP BY expense_type
        ORDER BY "Avg Days to Pay" ASC
    """
    return pd.read_sql_query(sql, conn)


def query_overdue_by_cost_center(conn):
    """
    Q3: Which cost centers carry the most overdue exposure?

    Business meaning:
    Cost center overdue exposure tells management which part of
    the business has the most unpaid liability outstanding.
    This supports month-end accrual decisions and helps budget
    owners understand their outstanding commitments.
    """
    sql = """
        SELECT
            cost_center                                     AS "Cost Center",
            COUNT(*)                                        AS "Overdue Invoices",
            ROUND(SUM(outstanding_amount_usd), 2)           AS "Overdue Exposure (USD)",
            ROUND(AVG(days_past_due), 1)                    AS "Avg Days Past Due",
            MAX(days_past_due)                              AS "Oldest (Days)"
        FROM ap_data
        WHERE is_overdue = 1
        GROUP BY cost_center
        ORDER BY "Overdue Exposure (USD)" DESC
    """
    return pd.read_sql_query(sql, conn)


def query_pareto_vendor_spend(conn):
    """
    Q4: Which vendors represent the top 80% of total spend? (Pareto)

    Business meaning:
    In most AP operations, 20% of vendors represent 80% of spend.
    Identifying this concentration helps management focus vendor
    relationship efforts, negotiate better payment terms, and
    assess supply chain risk from over-reliance on key suppliers.
    """
    sql = """
        SELECT
            vendor_name                                             AS "Vendor",
            ROUND(SUM(invoice_amount_usd), 2)                      AS "Total Spend (USD)",
            ROUND(
                100.0 * SUM(invoice_amount_usd) /
                SUM(SUM(invoice_amount_usd)) OVER (), 2
            )                                                       AS "Spend %",
            ROUND(
                100.0 * SUM(SUM(invoice_amount_usd)) OVER (
                    ORDER BY SUM(invoice_amount_usd) DESC
                ) /
                SUM(SUM(invoice_amount_usd)) OVER (), 2
            )                                                       AS "Cumulative %"
        FROM ap_data
        GROUP BY vendor_name
        ORDER BY "Total Spend (USD)" DESC
    """
    df = pd.read_sql_query(sql, conn)
    df["In Top 80%"] = df["Cumulative %"].apply(
        lambda x: "Yes" if x <= 80 else "No"
    )
    return df


def query_missed_discount_by_vendor(conn):
    """
    Q5: How much discount value was missed by vendor?

    Business meaning:
    Missed early payment discounts are a direct cost to the business.
    Breaking this down by vendor shows which relationships are costing
    the most in unrealized savings and which vendors AP should
    prioritize for faster payment processing.
    """
    sql = """
        SELECT
            vendor_name                                             AS "Vendor",
            COUNT(*)                                                AS "Eligible Invoices",
            SUM(CASE WHEN discount_captured_calc = 1 THEN 1 ELSE 0 END) AS "Captured",
            SUM(CASE WHEN discount_missed_calc = 1 THEN 1 ELSE 0 END)   AS "Missed",
            ROUND(SUM(captured_discount_value), 2)                  AS "Value Captured (USD)",
            ROUND(SUM(missed_discount_value), 2)                    AS "Value Missed (USD)"
        FROM ap_data
        WHERE is_discount_eligible = 1
        GROUP BY vendor_name
        ORDER BY "Value Missed (USD)" DESC
    """
    return pd.read_sql_query(sql, conn)


def query_on_time_rate_by_payment_method(conn):
    """
    Q6: What is the on-time payment rate by payment method?

    Business meaning:
    ACH, Wire, and Check have different processing times and costs.
    This query shows whether the payment method chosen affects
    on-time performance — useful for optimizing payment method
    selection by vendor type and invoice urgency.
    """
    sql = """
        SELECT
            payment_method                                                  AS "Payment Method",
            COUNT(*)                                                        AS "Total Invoices",
            SUM(CASE WHEN late_payment_flag = 0 THEN 1 ELSE 0 END)         AS "On Time",
            SUM(CASE WHEN late_payment_flag = 1 THEN 1 ELSE 0 END)         AS "Late",
            ROUND(
                100.0 * SUM(CASE WHEN late_payment_flag = 0 THEN 1 ELSE 0 END)
                / COUNT(*), 1
            )                                                               AS "On-Time Rate %",
            ROUND(AVG(days_to_pay), 1)                                      AS "Avg Days to Pay"
        FROM ap_data
        WHERE status = 'Paid'
            AND payment_method IS NOT NULL
            AND payment_method != ''
        GROUP BY payment_method
        ORDER BY "On-Time Rate %" DESC
    """
    return pd.read_sql_query(sql, conn)


def query_longest_blocked_invoices(conn):
    """
    Q7: Which invoices have been blocked the longest?

    Business meaning:
    Blocked invoices represent a workflow failure — something
    prevented the invoice from being approved for payment.
    The longer a block sits unresolved, the higher the risk of
    vendor escalation, late fees, or supply disruption.
    This query surfaces the most urgent items for immediate action.
    """
    sql = """
        SELECT
            invoice_num                                 AS "Invoice",
            vendor_name                                 AS "Vendor",
            expense_type                                AS "Expense Type",
            ROUND(invoice_amount_usd, 2)                AS "Amount (USD)",
            posting_date                                AS "Posted",
            days_open_as_of_snapshot                    AS "Days Blocked",
            urgency_level                               AS "Urgency"
        FROM ap_data
        WHERE status = 'Blocked'
        ORDER BY days_open_as_of_snapshot DESC
    """
    return pd.read_sql_query(sql, conn)


def query_exception_summary(conn):
    """
    Q8: Exception summary by type and total exposure.

    Business meaning:
    This is the exception report a controller reviews at month end.
    It shows the count and dollar exposure of every exception category
    so management can prioritize resolution by financial impact.
    Sorted by urgency — Red items always appear first.
    """
    sql = """
        SELECT
            issue_type                                  AS "Issue Type",
            urgency_level                               AS "Urgency",
            COUNT(*)                                    AS "Count",
            ROUND(SUM(outstanding_amount_usd), 2)       AS "Total Exposure (USD)",
            ROUND(AVG(days_open_as_of_snapshot), 1)     AS "Avg Days Open"
        FROM ap_data
        WHERE is_exception = 1
            AND issue_type != ''
            AND issue_type IS NOT NULL
        GROUP BY issue_type, urgency_level
        ORDER BY
            CASE urgency_level
                WHEN 'Red'    THEN 1
                WHEN 'Amber'  THEN 2
                WHEN 'Yellow' THEN 3
                ELSE 4
            END,
            "Total Exposure (USD)" DESC
    """
    return pd.read_sql_query(sql, conn)


# ─────────────────────────────────────────────
# EXCEL EXPORT — FORMATTED OUTPUT
# ─────────────────────────────────────────────

def export_to_excel(results, output_path):
    """
    Export all query results to a formatted Excel workbook.
    Each query gets its own tab with a header, column formatting,
    and color-coded rows where urgency data is present.

    Business meaning:
    Printing results to terminal is useful for developers.
    A formatted Excel file is useful for controllers, managers,
    and anyone who needs to share findings in a meeting or email.
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, (title, df) in results.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)

    # Now apply formatting
    wb = load_workbook(output_path)

    for sheet_name, (title, df) in results.items():
        ws = wb[sheet_name]

        # ── Title row ──
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(df.columns)
        )
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(
            name="Aptos Narrow", bold=True, size=12, color=WHITE
        )
        title_cell.fill = PatternFill("solid", fgColor=NAVY)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── Column header row (row 3) ──
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = Font(
                name="Aptos Narrow", bold=True, size=10, color=WHITE
            )
            cell.fill = PatternFill("solid", fgColor=STEEL_BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                bottom=Side(style="thin", color=GOLD)
            )
        ws.row_dimensions[3].height = 20

        # ── Data rows ──
        for row_idx in range(4, len(df) + 4):
            excel_row = row_idx
            row_fill = WHITE if (row_idx % 2 == 0) else LIGHT_GRAY

            # Check urgency column for color coding
            urgency_col = None
            for col_idx, col_name in enumerate(df.columns, start=1):
                if col_name == "Urgency":
                    urgency_col = col_idx
                    break

            urgency_value = None
            if urgency_col:
                urgency_cell = ws.cell(row=excel_row, column=urgency_col)
                urgency_value = urgency_cell.value

            if urgency_value == "Red":
                row_fill = "FFE0E0"
            elif urgency_value == "Amber":
                row_fill = "FFF0E0"
            elif urgency_value == "Yellow":
                row_fill = "FFFDE0"

            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.fill = PatternFill("solid", fgColor=row_fill)
                cell.font = Font(name="Aptos Narrow", size=10)
                cell.alignment = Alignment(vertical="center")

        # ── Column widths ──
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_width = max(
                len(str(col_name)),
                df.iloc[:, col_idx - 1].astype(str).str.len().max()
                if not df.empty else 0
            )
            ws.column_dimensions[col_letter].width = min(max_width + 4, 40)

        # ── Freeze header rows ──
        ws.freeze_panes = "A4"

        # ── Hide gridlines ──
        ws.sheet_view.showGridLines = False

    wb.save(output_path)


# ─────────────────────────────────────────────
# MAIN WORKFLOW
# ─────────────────────────────────────────────

def main():
    """
    Main workflow:
    1. Validate input file exists
    2. Load data into SQLite with duplicate check
    3. Run all 8 SQL analysis queries
    4. Export formatted results to Excel
    5. Print summary to terminal
    """
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        raise FileNotFoundError(
            f"\n  Input file not found: {INPUT_FILE}\n"
            f"  Run ap_dashboard_backend.py first to generate this file.\n"
        )

    print("\n" + "=" * 65)
    print("  AP AD-HOC ANALYSIS — Snapshot: June 30, 2025")
    print("  Powered by Python + SQLite")
    print("=" * 65)

    conn, df = load_data_to_sqlite(INPUT_FILE)
    print(f"  Data loaded: {len(df)} invoices ready for analysis")
    print("=" * 65)

    # Run all queries and collect results
    # Format: { tab_name: (tab_title, dataframe) }
    results = {
        "Q1 Late Payments":     (
            "Q1 — Which vendors have the highest late payment frequency?",
            query_late_payment_frequency(conn)
        ),
        "Q2 Days by Expense":   (
            "Q2 — Average days to pay by expense type",
            query_avg_days_to_pay_by_expense(conn)
        ),
        "Q3 Overdue by CC":     (
            "Q3 — Overdue exposure by cost center",
            query_overdue_by_cost_center(conn)
        ),
        "Q4 Pareto Spend":      (
            "Q4 — Pareto: which vendors represent the top 80% of spend?",
            query_pareto_vendor_spend(conn)
        ),
        "Q5 Missed Discounts":  (
            "Q5 — Missed discount value by vendor",
            query_missed_discount_by_vendor(conn)
        ),
        "Q6 Payment Methods":   (
            "Q6 — On-time payment rate by payment method",
            query_on_time_rate_by_payment_method(conn)
        ),
        "Q7 Blocked Invoices":  (
            "Q7 — Longest blocked invoices requiring immediate attention",
            query_longest_blocked_invoices(conn)
        ),
        "Q8 Exception Summary": (
            "Q8 — Exception summary by type and total exposure",
            query_exception_summary(conn)
        ),
    }

    # Print row counts to terminal so user knows queries ran
    for tab_name, (title, result_df) in results.items():
        print(f"  {tab_name:<25} {len(result_df)} rows")

    conn.close()

    # Export to Excel
    export_to_excel(results, OUTPUT_FILE)

    print("=" * 65)
    print(f"  Results exported to: {OUTPUT_FILE}")
    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()
